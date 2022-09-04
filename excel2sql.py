#! /usr/bin/env python3

'''
This code was based off of: https://gist.github.com/antiproblemist/0c2694cc17d7e39e9d12
License: MIT
Maintainer: Ross Miller <rossnmiller@gmail.com>
Example usage:
  ./scripts/excel2sql.py --debug false --user root --password password --host 127.0.0.1 --database dbname ./Sheet.xlsx

This code uses the openpyxl package for playing around with excel using Python code
to convert complete excel workbook (all sheets) to an SQL database
The code assumes that the first row of every sheet is the column name
Every sheet is stored in a separate table
The sheet name is assigned as the table name for every sheet

This code will assume that the first row in each sheet is the list of column names (which match the order and number of columns in the database for that respective table)

This code will assume that any subsequent row that has a colored background should be ignored. This is useful for having an "example" row in each table that shows what a valid row looks like.
'''

import os
import sys
import pymysql
import argparse
import openpyxl
from openpyxl import load_workbook
import re

pymysql.install_as_MySQLdb()

# # suppress annoying mysql warnings
# warnings.filterwarnings(action='ignore', category=pymysql.Warning)

def main(debug, input_file, user, password, host, database):
	is_debug_mode = debug == "true"
	if is_debug_mode:
		print("----- DEBUG MODE -----")

	db = pymysql.connect(host=host, user=user, passwd=password)
	cursor = db.cursor()
	# create database and if doesn't exist
	cursor.execute('CREATE DATABASE IF NOT EXISTS %s;' % database)
	db.select_db(database)

	#replace with the complete path to youe excel workbook
	wb = load_workbook(filename=input_file)

	sheets = wb.sheetnames

	for sheet in sheets:
		print("\nSheet name: " + sheet + "\n")
		ws = wb[sheet]

		columns= []
		query = 'CREATE TABLE IF NOT EXISTS ' + str(sheet) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'

		for rows in ws.rows:
			for cell in rows:
				# print('  Found Column %s %s' % (cell.coordinate,cell.value))

				query += ', ' + cell.value + ' TEXT'
				columns.append(cell.value)
			break

		query += ');'
		print(query)
		# cursor.execute(query)
		# print('\n')

		type_formats = []

		tup = []
		for i, rows in enumerate(ws):
			tuprow = []

			if i != 0: # Skip the first row
				for row in rows:
					if rows[0].fill.start_color.rgb == "0000000": # Ignore lines that have a colored background
						# print(type(row.value))
						# print(row.value)
						if isinstance(row.value, str):
							type_formats.append("%s")
							# print("Detected string")

							row_value = ""

							# If string looks like "=Venues!A2", then it's an equation that needs to be resolved!
							if row.value[0] == '=':
								###
								#  Reference: https://openpyxl.readthedocs.io/en/stable/formula.html
								#  Example: This line below references cell "A2" in the sheet named "Clients"
								#   "=Clients!A2"
								#  The code block below is responsible for converting "=Clients!A2" into the value of cell "A2" in the "Clients" sheet
								###
								split_string = (row.value[1:]).split("!")
								referenced_sheet_name = split_string[0]
								referenced_cell = split_string[1]

								resolved_cell = wb[referenced_sheet_name][referenced_cell]
								row_value = resolved_cell.value
							else:
								row_value = str(row.value).strip()

							tuprow.append(row_value)
						elif isinstance(row.value, int):
							type_formats.append("%s")
							# print("Detected int")
							tuprow.append(int(row.value))
						elif isinstance(row.value, float):
							type_formats.append("%s")
							if(row.value / 10 * 10 == row.value):
								# print("Detected int")
								tuprow.append(int(row.value))
							else:
								# print("Detected float")
								tuprow.append(float(row.value))
						else:
							type_formats.append("%s")
							# print("Detected unknown")
							tuprow.append(row.value)

				tup.append(tuple(tuprow))

		insQuery1 = 'INSERT INTO ' + str(sheet) + ' ('
		insQuery2 = ''
		for col in columns:
			insQuery1 += col + ', '
		for type_format in type_formats:
			insQuery2 += type_format + ', '
		insQuery1 = insQuery1[:-2] + ') VALUES ('
		insQuery2 = insQuery2[:-2] + ')'
		insQuery = insQuery1 + insQuery2

		if is_debug_mode == False:
			cursor.executemany(insQuery, tup)

		print("")
		print(insQuery)
		print("")
		for tup_item in tup:
			print(tup_item)
		print("\n-------")

		db.commit()
	db.close()

if __name__ == '__main__':
	parser = argparse.ArgumentParser(description='Convert Excel file with multiple sheets into SQT create & insert statements')
	parser.add_argument('--debug', dest='debug', default='false', help='If "true", does not execute SQL queries, just prints them out')
	parser.add_argument('--database', dest='database', default='test', help='Set the name of the database')
	parser.add_argument('--user', dest='user', default='root', help='The MySQL login username')
	parser.add_argument('--password', dest='password', default='', help='The MySQL login password')
	parser.add_argument('--host', dest='host', default='localhost', help='The MySQL host')
	parser.add_argument('input_file', help='The input excel file (in .xls or .xlsx format)')
	args = parser.parse_args(sys.argv[1:])

	main(args.debug, args.input_file, args.user, args.password, args.host, args.database)