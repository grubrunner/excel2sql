#! /usr/bin/env python3

'''
License: MIT
Maintainer: Ross Miller <rossnmiller@gmail.com>
Example usage:
  ./scripts/excel2sql.py --debug false --user root --password password --host 127.0.0.1 --database dbname ./Sheet.xlsx

This code creates a blank excel document out of a SQL database. Each table in the SQL database will turn into a sheet in the excel document. Each column in each table will turn into a column in the corresponding excel sheet
'''

import os
import sys
import pymysql
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

pymysql.install_as_MySQLdb()

# This supports up to 26 * 3 columns. To add more columns, update this list
valid_column_identifier_names = [
	"A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
	"AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
	"BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
]

fill_color = PatternFill(fill_type="solid", start_color='FFC7EFF0', end_color='FFC7EFF0')


# # suppress annoying mysql warnings
# warnings.filterwarnings(action='ignore', category=pymysql.Warning)

wb = Workbook()

def main(debug, output_file, table_blacklist, first_table_names, include_example_row, user, password, host, database):
	table_blacklist_as_list = table_blacklist.split(',')
	first_table_names_as_list = first_table_names.split(',')

	is_example_row_included = include_example_row == "true"

	is_debug_mode = debug == "true"
	if is_debug_mode:
		print("----- DEBUG MODE -----")

	db = pymysql.connect(host=host, user=user, password=password, db=database)

	cursor = db.cursor()

	cursor.execute("SHOW tables")

	result = cursor.fetchall()

	table_names = first_table_names_as_list

	for i in range(len(result)):
		table_name = result[i][0]

		if table_name not in table_names:
			table_names.append(table_name)

	for table_name in table_names:
		if table_name == "":
			continue;
		if table_name not in table_blacklist_as_list:
			cursor2 = db.cursor()
			cursor2.execute("DESCRIBE " + table_name)
			column_results = cursor2.fetchall()
			if is_debug_mode:
				print("\n\n-----\n"+ table_name)

			worksheet = wb.create_sheet(table_name)

			first_column_name = ""

			for column_index in range(len(column_results)):
				column_name = column_results[column_index][0]

				if column_index == 0:
					first_column_name = column_name

				if is_debug_mode:
					print("  " + column_name)

				worksheet[valid_column_identifier_names[column_index] + "1"] = column_name

			if is_example_row_included:
				cursor2.execute("SELECT * FROM " + table_name + " ORDER BY " + first_column_name + " DESC LIMIT 1")
				last_row_results = cursor2.fetchall()

				last_row_results_list = list(last_row_results)

				# print("last_row_results len " + str())

				if len((last_row_results_list)) == 1:
					last_row_columns = list(last_row_results_list[0])

					for last_row_columns_index in range(len(last_row_columns)):

						last_row_results_column_name = last_row_columns[last_row_columns_index]

						# print("last_row_results_column_name type " + str(type(last_row_results_column_name)))

						worksheet[valid_column_identifier_names[last_row_columns_index] + "2"] = last_row_results_column_name

						worksheet[valid_column_identifier_names[last_row_columns_index] + "2"].fill = fill_color

	del wb['Sheet'] # Delete the first sheet that was auto created my openpyxl
	wb.save(output_file)

if __name__ == '__main__':
	parser = argparse.ArgumentParser(description='Convert Excel file with multiple sheets into SQT create & insert statements')
	parser.add_argument('--debug', dest='debug', default='false', help='If "true", does not execute SQL queries, just prints them out')
	parser.add_argument('--database', dest='database', default='test', help='Set the name of the database')
	parser.add_argument('--user', dest='user', default='root', help='The MySQL login username')
	parser.add_argument('--password', dest='password', default='', help='The MySQL login password')
	parser.add_argument('--host', dest='host', default='localhost', help='The MySQL host')
	parser.add_argument('--table_blacklist', dest='table_blacklist', default='', help='A comma seperated list (with no spaces) of table names that should not be included in the excel spreadsheet')
	parser.add_argument('--first_table_names', dest='first_table_names', default='', help='A comma seperated list of table namaes that need to be filled out first. If a table is not included in this list, then the order will be alphabetical')
	parser.add_argument('--include_example_row', dest='include_example_row', default='false', help='If set to "true", the first row will have a colored background and will include the result from the last row in the database for the respective table as an example of what a valid row in this table looks like')
	parser.add_argument('output_file', help='The output excel file (in .xls or .xlsx format)')
	args = parser.parse_args(sys.argv[1:])

	main(args.debug, args.output_file, args.table_blacklist, args.first_table_names, args.include_example_row, args.user, args.password, args.host, args.database)